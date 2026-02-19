import requests
from azure.identity import AzureCliCredential
from datetime import datetime, timedelta, UTC
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.chart import BarChart, Reference
from collections import Counter

# -------------------------------
# CONFIGURATION
# -------------------------------
EXCEL_FILENAME = "expiring_secrets_report.xlsx"
DAYS_THRESHOLD = 30  # alert window

# Priority colors
PRIORITY_COLORS = {
    "0-10": "FF0000",   # Red → urgent
    "11-20": "FFA500",  # Orange → medium
    "21-30": "FFFF00",  # Yellow → low
}

# -------------------------------
# AUTHENTICATION
# -------------------------------
credential = AzureCliCredential()
token = credential.get_token("https://graph.microsoft.com/.default")
headers = {"Authorization": f"Bearer {token.token}"}

# -------------------------------
# FETCH APPLICATIONS
# -------------------------------
url = "https://graph.microsoft.com/v1.0/applications?$select=displayName,appId,passwordCredentials"
threshold = datetime.now(UTC) + timedelta(days=DAYS_THRESHOLD)

expiring_secrets = []

while url:
    response = requests.get(url, headers=headers)
    response.raise_for_status()
    data = response.json()

    for app in data.get("value", []):
        display_name = app.get("displayName")
        app_id = app.get("appId")

        for cred in app.get("passwordCredentials", []):
            end_date = cred.get("endDateTime")
            if end_date:
                expiry = datetime.fromisoformat(end_date.replace("Z", "+00:00"))
                days_left = (expiry - datetime.now(UTC)).days
                if 0 <= days_left <= DAYS_THRESHOLD:
                    # Assign priority bucket
                    if days_left <= 10:
                        priority = "0-10"
                    elif days_left <= 20:
                        priority = "11-20"
                    else:
                        priority = "21-30"

                    expiring_secrets.append({
                        "Display Name": display_name,
                        "Client ID": app_id,
                        "Expires On": expiry.strftime('%Y-%m-%d'),
                        "Days Left": days_left,
                        "Priority": priority
                    })

    url = data.get("@odata.nextLink")

# -------------------------------
# WRITE TO EXCEL
# -------------------------------
wb = Workbook()
ws = wb.active
ws.title = "Expiring Secrets"

# Headers
headers = ["Display Name", "Client ID", "Expires On", "Days Left", "Priority"]
ws.append(headers)

# Fill rows with data & colors
for secret in expiring_secrets:
    ws.append([secret[h] for h in headers])
    row = ws.max_row
    color = PRIORITY_COLORS.get(secret["Priority"], "FFFFFF")
    for col in range(1, len(headers)+1):
        ws.cell(row=row, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

# -------------------------------
# CREATE SUMMARY TABLE FOR CHART
# -------------------------------
# Count number of secrets per priority
priority_counts = Counter([s["Priority"] for s in expiring_secrets])
summary_ws = wb.create_sheet(title="Summary")
summary_ws.append(["Priority", "Number of Secrets"])
for priority in ["0-10", "11-20", "21-30"]:
    summary_ws.append([priority, priority_counts.get(priority, 0)])

# -------------------------------
# CREATE BAR CHART
# -------------------------------
chart = BarChart()
chart.title = "Expiring Secrets Priority"
chart.y_axis.title = "Number of Secrets"
chart.x_axis.title = "Priority (Days Left)"

data = Reference(summary_ws, min_col=2, min_row=1, max_row=4)
categories = Reference(summary_ws, min_col=1, min_row=2, max_row=4)
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)
chart.shape = 4
summary_ws.add_chart(chart, "D2")

# -------------------------------
# SAVE FILE
# -------------------------------
wb.save(EXCEL_FILENAME)
print(f"Excel report generated: {EXCEL_FILENAME} with Pivot-like summary and chart.")
